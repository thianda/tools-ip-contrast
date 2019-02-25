#!/usr/bin/python3
# -*- coding: utf-8 -*-

import platform
import configparser
import os
import xlrd
from datetime import datetime
import openpyxl
import re
import traceback
# import shutil

__version__ = '0.8.3'
configFileName = 'config_%s.ini' % __version__
DEBUG_FILE = 'debug_log.txt'
config = configparser.ConfigParser()


# 当前时间的字符串
def now():
    return datetime.strftime(datetime.now(), '%Y-%m-%d %H:%M:%S')


# ip 转换： int 转 str
def int2ip(x):
    return '.'.join([str(int(x / (256 ** i) % 256)) for i in range(3, -1, -1)])


# ip 转换： str 转 int
def ip2int(x):
    try:
        return sum([256 ** j * int(i) for j, i in enumerate(x.split('.')[::-1])])
    except:
        return False


# ip 字符串 10.10.10.10/32 转成数组: ip/mask/ip_start/ip_end
def ipParse(ipStr):
    if ipStr == '':
        return [None, None, None, None]
    maskLen = 32
    if '/' in ipStr:
        ipStr, maskLen = ipStr.split('/')
    # print(ipStr, maskLen)
    ipInt = ip2int(ipStr)
    if maskLen == 32:
        return [ipInt, 4294967295, ipInt, ipInt]
    mask = 0xFFFFFFFF << (32 - int(maskLen))
    # print('mask:', mask)
    ipStart = ipInt & mask & 0xFFFFFFFF
    ipEnd = ipInt | ~mask & 0xFFFFFFFF
    return [ipInt, mask, ipStart, ipEnd]


# ip, mask 转成 字符串 10.10.10.10/32
def ipImport(ipInt, mask):
    pass


# 检查配置格式是否正确
def checkConfig(configFileName):
    global config
    config.read(configFileName)
    # _version = config.get('common', 'version', fallback=None)
    # if not _version:
    #     return False
    # _version = [int(i) for i in _version.split('.')]
    # _current_ver = [int(i) for i in __version__.split('.')]
    # if _version < _current_ver:
    #     return False
    sections = config.sections()
    sectionLens = len(sections)
    if sectionLens < 5:
        return False
    # 每个 section 中的 options 的个数必须是相同的
    SETTING_LENS = 2
    for i in range(SETTING_LENS, sectionLens):
        if i < sectionLens - SETTING_LENS:
            n1 = len(config.options(sections[i]))
            n2 = len(config.options(sections[i + 1]))
            if n1 == 0 | n1 != n2:
                return False
    # 可进一步判断 options 名是否一致
    pass
    return True


# 释放默认配置
def writeConfig(configFileName):
    global config
    config['common'] = {
        'version': __version__
    }
    config['对比文件名'] = {
        '省内资管': 'IP地址.',
        '集团': '-IP地址-',
        '工信部备案': 'fpxxList',
    }
    config['省内资管'] = {
        'tag1': '所属地市',
        'ip': 'IP地址*',
        'starttime': '分配使用时间',
        'field3': '联系人姓名(客户侧)',
        'field4': '联系电话(客户侧)',
        'field5': '单位详细地址',
        'field6': '联系人邮箱(客户侧)',
        'field7': '单位名称/具体业务信息',
    }
    config['集团'] = {
        'tag1': '所属地市',
        'ip': '网段名称',
        'starttime': '分配使用时间',
        'field3': '联系人姓名(客户侧)',
        'field4': '联系人电话(客户侧)',
        'field5': '单位详细地址',
        'field6': '联系人邮箱(客户侧)',
        'field7': '单位名称/具体业务信息',
    }
    config['工信部备案'] = {
        'tag1': '所属地',
        'ip': '起始IP;终止IP',
        'starttime': '分配日期',
        'field3': '联系人姓名',
        'field4': '联系人电话',
        'field5': '单位详细地址',
        'field6': '联系人电子邮件',
        'field7': '使用单位名称',
    }
    with open(configFileName, 'w') as configFile:
        configFile.write(
            '''# Author: Xianda

# 本配置为一致性检查工具的配置
# 如删除本配置文件，重新生成的配置文件即为默认配置
# 如需修改配置，修改本文件后直接保存即可

######

# before 表示数据的起始行（列名占用的行数）。现已自动识别（配置无效）
# 若 IP 地址字段分为起始IP、终止IP的，`ip` 字段中用“;”(英文分号)隔开
# 程序会依次对 field 字段进行对比并输出对比结果
# field 字段有变化可直接在此增删改，满足一一对应即可



''')
        config.write(configFile)


# 初始化配置
def initConfig():
    global configFileName
    if os.path.exists(configFileName):
        if not checkConfig(configFileName):
            print('Warning:', '配置有误，已恢复默认', configFileName)
    else:
        print('Info:', '配置已更新', configFileName)
        # 释放默认配置，开始对比数据
        writeConfig(configFileName)
    print('当前加载的配置：', configFileName)
    return contrast()


# 读取配置
def readConfig(section, option):
    global config
    global configFileName
    config.read(configFileName)
    value = config.get(section, option)
    return value


# 匹配导出数据文件名
def matchedFileName():
    global config
    global configFileName
    fileName = {}
    config.read(configFileName)
    for option in config.options('对比文件名'):
        value = config.get('对比文件名', option)
        fileName[option] = []
        for f in os.listdir():
            if value in f and not '~' in f:
                fileName[option].append(f)
        if not fileName[option]:
            fileName.pop(option)
            print('Warning:', '未找到', option, '的导出数据，将不进行该数据的一致性检查。')
    return fileName


# 生成中间文件，转换导出数据为每一行一个 ip
def generateTemp(fileName):
    if fileName == {}:
        print('Error:', '未识别到excel文件。')
        pause()
        exit()
    global config
    # print(list(fileName.keys())[0])
    # print(config.options(list(fileName.keys())[0]))

    # 基于 ip 列 生成中间文件，并补充 ipStart、ipEnd 列

    # 智能识别、生成配置、并输出中间文件，在一次 fileName 的 for 循环中完成
    options = {}
    colNames = {}
    ipNames = {}
    sheet = {}
    # 按导出文件名遍历
    _t = datetime.now()
    print(now(), '识别表格配置...')
    for k, vv in fileName.items():
        # 遍历同一类型的所有文件，如集团导出文件分多个
        _totalRows = 0
        for v in vv:
            # 添加计算进度百分百以及预计用时的计算
            pass
        # 获取第一个文件的配置
        v = vv[0]
        xls = xlrd.open_workbook(v)
        # 默认获取第一个 sheet 的第一行做为列名所在的行
        Row0 = xls.sheet_by_index(0).row_values(0)
        options[k] = {}
        options[k]['fieldCols'] = {}
        options[k]['ipCols'] = []
        # options[k]['output'] = []
        colNames[k] = []
        ipNames[k] = []
        fields = config.options(k)
        # 遍历配置文件中要对比的列名 记录要对比的字段所在的列
        for field in fields:
            configValue = config.get(k, field)
            if 'before' == field:
                # 数据起始行（标题所占的行数）默认为 1
                options[k]['before'] = configValue
            else:
                colFounded = False
                # 遍历导出数据的第一行单元格
                for i in range(0, len(Row0)):
                    cellValue = str(Row0[i].strip())
                    if cellValue in configValue:
                        colFounded = True
                        if 'ip' == field:
                            # 记录 ip 所在的列
                            options[k]['ipCols'].append(i)
                            ipNames[k].append(cellValue)
                        else:
                            # 记录要对比的字段所在的列
                            options[k]['fieldCols'][field] = i
                            colNames[k].append(cellValue)
                        # 此处不能加 break，否则匹配到`起始IP`即跳出 for 循环，无法匹配`结束IP`
                # 若根据配置未找到对应的列：给出提示，结束运行
                if not colFounded:
                    print('在%s中，未找到数据列：【%s】，工具即将退出。' % (k, configValue))
                    pause()
                    exit()
        # colNames[k] = ['ipStart', 'ipEnd'] + colNames[k]
        colNames[k].extend(['ipStart', 'ipEnd'])
        colNames[k].extend(ipNames[k])
        # del(xls, Row0, field, v)
    t = (datetime.now() - _t).total_seconds()
    print(now(), '表格配置识别完毕。用时 %2.4f 秒。' % t)
    # print('Info:', 'options', options, '\n')
    # print('Info:', 'colNames', colNames, '\n')
    # os.mkdir('\\_temp')
    SEP = os.path.sep
    configPath = os.getcwd() + SEP + '..' + SEP + '__output' + SEP
    # tempdir=>C:\\ProgramData
    # tempdir = os.environ.get('ALLUSERSPROFILE')
    # tempdir=>%USERPROFILE%/Local/Temp, C:/Users/XXXX/AppData/Local/Temp
    tempdir = os.environ.get('TEMP')
    if tempdir and os.path.exists(tempdir):
        configPath = tempdir + SEP + 'Xianda' + SEP + 'ipContrast' + SEP
    if not os.path.exists(configPath):
        os.makedirs(configPath)
    # 创建中间文件
    note = '对比结果在第二个sheet页最后 ' + str(len(colNames) - 1) + ' 列'
    author = '--Xianda'
    # wb = openpyxl.Workbook(write_only=True)
    # ws0 = wb.create_sheet('说明')
    # from openpyxl.worksheet.write_only import WriteOnlyCell
    # from openpyxl.styles import Font
    # cell = WriteOnlyCell(ws0, value=note)
    # cell.font = Font(size=18, color='FF0000')
    # ws0.append([None])
    # ws0.append([None, cell])
    # ws0.append([None])
    # cell = WriteOnlyCell(ws0, value=author)
    # cell.font = Font(size=11, color='8060ee', bold=True)
    # ws0.append([None, None, cell])
    # cell = WriteOnlyCell(ws0, value=str(options))
    # cell.font=Font(size=11, color='666600')
    # ws0.append([None, cell])
    # cell=WriteOnlyCell(ws0, value=str(colNames))
    # cell.font=Font(size=11, color='666600')
    # ws0.append([None, cell])
    wb = openpyxl.Workbook()
    ws0 = wb.active  # 第一个 sheet
    ws0.title = '说明'
    ws0.column_dimensions['B'].width = 120
    ws0['B2'] = note
    ws0['C3'] = author
    line = 6
    count = len(fileName) + 1
    for i in fileName:
        ws0['A' + str(line)] = i
        ws0['B' + str(line)] = str(options[i])
        ws0['A' + str(line + count)] = i
        ws0['B' + str(line + count)] = str(colNames[i])
        line += 1
    # ws.merge_cells('A2:G3')
    del(ws0)
    currDate = datetime.strftime(datetime.now(), '%Y-%m-%d_%H%M%S')
    wrName = configPath + 'results-' + currDate + '.xlsx'
    isFirstSheet = True
    ipv4_reg = r'(25[0-5]|2[0-4]\d|[0-1]?\d?\d)(\.(25[0-5]|2[0-4]\d|[0-1]?\d?\d)){3}'
    for k, vv in fileName.items():
        print('\n%s 开始筛选数据：%s' % (now(), k))
        # 为每个对比文件创建一个 sheet
        ws = wb.create_sheet(k)
        if isFirstSheet:
            extendTitle = ['与' + x + '一致' for x in fileName.keys()]
            extendTitle.pop(0)
            title = colNames[k] + ['预留1', '所属地市', 'concatenate'] + extendTitle
        else:
            title = colNames[k]
        ws.append(title)
        ipCols = options[k]['ipCols']
        fieldCols = options[k]['fieldCols']
        # 遍历每个文件
        for v in vv:
            _t = datetime.now()
            xls = xlrd.open_workbook(v)
            t = (datetime.now() - _t).total_seconds()
            print(now(), '加载文件 %s 用时 %2.4f 秒。解析中...' % (v, t))
            # 获取所有 sheet
            sheet[k] = [xls.sheet_by_index(_index)
                        for _index in range(0, len(xls.sheet_names()))]
            # 遍历每个 sheet
            sheets_len = len(sheet[k])
            _nrows = 0
            for i_sheet in range(0, sheets_len):
                _sheet = sheet[k][i_sheet]
                if 'Sheet' in _sheet.name:
                    # 疑似修改导出文件手动创建的 sheet，跳过
                    continue
                # 自动识别当前 sheet 的 before
                _invalid = True  # 未识别到有效数据为无效
                _before = 0
                # _before = int(options[k]['before'])
                for i in range(10):
                    col_ip = _sheet.row_values(i)[ipCols[0]]
                    # if i < 2:
                    #     print(_sheet.row_values(i))
                    #     printYellow(col_ip)
                    _before += 1
                    is_ipv4 = re.search(ipv4_reg, col_ip)
                    if is_ipv4 or ':' in col_ip:
                        _invalid = False
                        break
                if _invalid:
                    print(now(), 'Error：未识别到有效数据，已跳过：')
                    printRed('%s => [%s]' % (v, _sheet.name))
                    continue
                # 遍历每一行数据
                _nrows += _sheet.nrows
                for row in range(_before, _nrows):
                    print('\r%s 解析进度：行数 %s/%s sheet %s/%s' %
                          (now(), row+1, _nrows, i_sheet+1, sheets_len), end='')
                    rowValues = _sheet.row_values(row)
                    ip1Str = rowValues[ipCols[0]]
                    if ':' in ip1Str:
                        # ipv6 暂不对比，过滤本行
                        continue
                    tempRow = []
                    if not isFirstSheet:
                        if rowValues[int(fieldCols['field5'])] == '':
                            # 过滤无用数据(field5 为空)
                            continue
                    currentRow = str(ws.max_row + 1)
                    for i in fieldCols:
                        cellValue = rowValues[int(fieldCols[i])]
                        # 转换 starttime 的格式
                        if i == 'starttime' and len(cellValue) > 10:
                            cellValue = cellValue[:10]
                        tempRow.append(cellValue)
                        # del(cellValue)
                    if len(ipCols) == 1:
                        ip2Str = None
                        if '/32' in ip1Str or not '/' in ip1Str:
                            ip_start = ip2int(ip1Str.split('/')[0])
                            ip_end = ip_start
                        else:
                            _ip_parse = ipParse(ip1Str)
                            ip_start = _ip_parse[2]
                            ip_end = _ip_parse[3]
                    elif len(ipCols) == 2:
                        # 导出数据自带起始 IP 结束IP
                        ip2Str = rowValues[ipCols[1]]
                        ip_start = ip2int(ip1Str)
                        ip_end = ip2int(ip2Str)
                        if ip_start > ip_end:
                            ip_start, ip_end = ip_end, ip_start
                            ip1Str, ip2Str = ip2Str, ip1Str
                    else:
                        # ip 字段个数不是 1 也不是 2
                        print('Warning:', 'IP 列识别错误 %s %s 行' % (k, currentRow))
                        return DEBUG_FILE
                    formula_strs = '=CONCATENATE(A%s,"-",B%s,"-",C%s,"-",D%s,"-",E%s,"-",F%s)' % (
                        (currentRow,)*6)
                    tempRow.extend(
                        [ip_start, ip_end, ip1Str, ip2Str, None, formula_strs])
                    if isFirstSheet:
                        formula = [x for x in fileName.keys()]
                        formula.pop(0)
                        formula_strs = ['=VLOOKUP(H%s,%s!G:L,6,0)=L%s' % (
                            currentRow, x, currentRow) for x in formula]
                        tempRow.extend(formula_strs)
                    # print('tempRow', tempRow) # debug
                    # 过滤客户地址为空的数据
                    # if tempRow[3] == '':
                    #     continue
                    # else:
                    ws.append(tempRow)
                    # del(tempRow, ip_start, ip_end, ip1Str, ip2Str, strings)
                    # gc.collect()
            ws.auto_filter.ref = 'A1:N' + str(ws.max_row)
            _t = datetime.now()
            wb.save(wrName)  # 每读取完一个文件保存一次
            _t = (datetime.now() - _t).total_seconds()
            print('\n%s 解析完毕 %s ' % (now(), v))
            print('当前进度已保存到： %s，保存操作用时 %2.4f 秒' % (wrName, _t))
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 11
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        isFirstSheet = False
        # ws.auto_filter.add_sort_condition('G2:G'+str(ws.max_row))
        # del(ws)
    wb.save(wrName)
    return wrName


def contrast():
    # 获取导出文件的文件名
    fileName = matchedFileName()
    # print('Info:','fileName', fileName, '\n')
    tempFile = generateTemp(fileName)
    return tempFile


# unused for now
# def copyfile(srcfile, dstfile):
#     if not os.path.isfile(srcfile):
#         print('%s not exist!' % (srcfile))
#     else:
#         fpath = os.path.split(dstfile)[0]  # 分离文件名和路径
#         if not os.path.exists(fpath):
#             os.makedirs(fpath)  # 创建路径
#         shutil.copyfile(srcfile, dstfile)  # 复制文件


# @测试 configparser
def _test_configparser():
    global config
    config.read(configFileName)
    sections = config.sections()
    for x in sections:
        options = config.options(x)
        for y in options:
            print(config.get(x, y))


if 'Windows' in platform.system():
    import sys
    import ctypes
    __stdInputHandle = -10
    __stdOutputHandle = -11
    __stdErrorHandle = -12
    __foreGroundBLUE = 0x09
    __foreGroundGREEN = 0x0a
    __foreGroundRED = 0x0c
    __foreGroundYELLOW = 0x0e
    stdOutHandle = ctypes.windll.kernel32.GetStdHandle(__stdOutputHandle)

    def setCmdColor(color, handle=stdOutHandle):
        return ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)

    def resetCmdColor():
        setCmdColor(__foreGroundRED | __foreGroundGREEN | __foreGroundBLUE)

    def printBlue(msg):
        setCmdColor(__foreGroundBLUE)
        sys.stdout.write(msg + '\n')
        resetCmdColor()

    def printGreen(msg):
        setCmdColor(__foreGroundGREEN)
        sys.stdout.write(msg + '\n')
        resetCmdColor()

    def printRed(msg):
        setCmdColor(__foreGroundRED)
        # sys.stdout.write(msg + '\n')
        print(msg)
        resetCmdColor()

    def printYellow(msg):
        setCmdColor(__foreGroundYELLOW)
        sys.stdout.write(msg + '\n')
        resetCmdColor()

    def pause():
        os.system('pause')

    def locateFile(file):
        print('\n- 即将跳转到输出结果 -\n')
        pause()
        os.system('explorer /select, ' + file)

else:
    STYLE = {
        'fore': {
            'red': 31,
            'green': 32,
            'yellow': 33,
            'blue': 34,
        }
    }

    def UseStyle(msg, mode='', fore='', back='40'):
        fore = '%s' % STYLE['fore'][fore] if STYLE['fore'].has_key(
            fore) else ''
        style = ';'.join([s for s in [mode, fore, back] if s])
        style = '\033[%sm' % style if style else ''
        end = '\033[%sm' % 0 if style else ''
        return '%s%s%s' % (style, msg, end)

    def printRed(msg):
        print(UseStyle(msg, fore='red'))

    def printGreen(msg):
        print(UseStyle(msg, fore='green'))

    def printYellow(msg):
        print(UseStyle(msg, fore='yellow'))

    def printBlue(msg):
        print(UseStyle(msg, fore='blue'))

    def pause():
        pass

    def locateFile(file):
        print('\n- 结果文件保存为：%s' % file)
        print('刷新页面可到 `__output` 文件夹中下载查看。')

if __name__ == '__main__':
    print('****欢迎使用一致性检查工具 %s\n' % __version__)
    t0 = datetime.now()
    print(now(), '**开始运行')
    result = DEBUG_FILE
    try:
        result = initConfig()
    except Exception as err:
        printRed('Error:')
        traceback.print_exc(file=open(DEBUG_FILE, 'w'))
        printRed(err)
        printBlue('出错啦。请反馈目录中的 debug_log.txt 文件内容')
    t = datetime.now() - t0
    print('\n执行用时：%2.4f s' % t.total_seconds())
    # os.system('pause')
    locateFile(result)
